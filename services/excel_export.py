# services/excel_export.py
import tempfile
from pathlib import Path
import pandas as pd
from aiogram.types import FSInputFile
from db import DBI
from config import RU_MONTHS
from utils.dates import human_ru_date
from openpyxl.styles import Alignment
from openpyxl import load_workbook

def _auto_width(ws, min_width: int = 10, max_width: int = 80, wrap: bool = False):

    """
    Auto-fit column widths by scanning headers + data (legacy helper).
    NOTE: The new logic below (_post_save_autofit) strictly follows the
    requirement: first save, then re-open and scan content-only.
    Kept for backward compatibility in case it's used elsewhere.
    """
    for col in ws.columns:
        try:
            letter = col[0].column_letter
        except Exception:
            continue

        max_len = 0
        for cell in list(col)[1:]:  # skip header row
            val = cell.value
            if val is None:
                ln = 0
            else:
                s = str(val)
                ln = max((len(line) for line in s.splitlines()), default=len(s))
            if ln > max_len:
                max_len = ln
            if wrap:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[letter].width = width

def _post_save_autofit(xlsx_path: Path, sheet_name: str, *, min_width: int = 10, max_width: int = 80, wrap: bool = True) -> None:
    """
    Strict post-save auto-fit by content only:
    - Column width = longest visible line length among rows 2..max_row * 1.2 (+2 padding), clamped to [min_width, max_width]
    - Optional wrap + vertical top for all data cells
    - Row height = base_height * max_lines_in_that_row (approx), so multiline is readable
    """
    print("ok")
    try:
        wb = load_workbook(filename=str(xlsx_path))
        ws = wb[sheet_name]
        print(f"Обрабатываем лист: {sheet_name}, строк: {ws.max_row}, столбцов: {ws.max_column}")

        header_cells = list(ws[1]) if ws.max_row >= 1 else []
        n_cols = len(header_cells)

        # First pass: compute max line length per column and max line count per row
        col_max_len: dict[int, int] = {c: 0 for c in range(1, n_cols + 1)}
        row_max_lines: dict[int, int] = {}

        for row_idx in range(2, ws.max_row + 1):
            max_lines_in_row = 1
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is None:
                    lines = [""]
                else:
                    s = str(val)
                    # normalize Windows/mac newlines to '\n'
                    s = s.replace("\r\n", "\n").replace("\r", "\n")
                    lines = s.split("\n")

                # longest visible line length in this cell
                longest = 0
                for line in lines:
                    # consider multiple spaces and tabs visually wider
                    l = len(line.expandtabs(4))
                    if l > longest:
                        longest = l

                if longest > col_max_len[col_idx]:
                    col_max_len[col_idx] = longest

                if len(lines) > max_lines_in_row:
                    max_lines_in_row = len(lines)

                if wrap:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            row_max_lines[row_idx] = max_lines_in_row

        # Second pass: set column widths
        for col_idx in range(1, n_cols + 1):
            # small scale factor to better approximate Excel width metrics
            scaled = int(col_max_len[col_idx]) + 1
            width = max(min_width, min(scaled, max_width))
            letter = ws.cell(row=1, column=col_idx).column_letter if n_cols else None
            if letter:
                ws.column_dimensions[letter].width = width

        # Third pass: set row heights (approx)
        base_height = 15  # Excel default row height is ~15 pt
        for row_idx, lines in row_max_lines.items():
            ws.row_dimensions[row_idx].height = base_height * max(1, lines)

        wb.save(str(xlsx_path))
        print("Автофит применён успешно")
    except Exception as e:
        print(f"Ошибка в _post_save_autofit: {e}")
        raise


def export_month_schedule(year: int, month: int) -> tuple[Path, int]:
    prefix = f"{year:04d}-{month:02d}-"
    with DBI._conn() as con:
        evs = con.execute("""
            SELECT date, COALESCE(type,''), COALESCE(title,''), COALESCE(time,''),
                   COALESCE(location,''), COALESCE(city,''), COALESCE(employee,''), COALESCE(duty_employee,''), COALESCE(info,'')
            FROM events WHERE date LIKE ? ORDER BY date, title, time
        """, (prefix+'%',)).fetchall()
    df = pd.DataFrame(evs, columns=["date","type","title","time","location","city","employee","duty_employee","info"])
    df["date"] = df["date"].map(human_ru_date)
    df = df.rename(columns={
        "date":"Дата","type":"Тип","title":"Название","time":"Время",
        "location":"Локация","city":"Город","employee":"Сотрудник","duty_employee":"Дежурный сотрудник","info":"Инфо"
    })
    out_path = Path(tempfile.gettempdir()) / f"График_{year}-{month:02d}.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet_name = "График"
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    # Post-save auto-fit strictly by content (skip headers)
    _post_save_autofit(out_path, sheet_name, min_width=10, max_width=80, wrap=True)
    return out_path, len(df)

def file_as_input(path: Path) -> FSInputFile:
    return FSInputFile(str(path))

def month_caption(year: int, month: int, updated: int) -> str:
    return f"График на {RU_MONTHS[month-1]} {year}. Обновлено назначений: {updated}"


# --- Export spectacles table ---
def export_spectacles_table() -> tuple[Path, int]:
    """
    Выгружает таблицу вида:
    Спектакль | Сотрудники (через запятую)
    """
    with DBI._conn() as con:
        rows = con.execute(
            """
            SELECT
                s.title AS title,
                COALESCE(GROUP_CONCAT(e.display, ', '), '') AS employees
            FROM spectacles s
            LEFT JOIN spectacle_employees se ON se.spectacle_id = s.id
            LEFT JOIN employees e          ON e.id = se.employee_id
            GROUP BY s.id, s.title
            ORDER BY s.title COLLATE NOCASE
            """
        ).fetchall()

    df = pd.DataFrame(rows, columns=["title", "employees"]).rename(columns={
        "title": "Спектакль",
        "employees": "Сотрудники",
    })
    # Make employees appear one per line for better readability and wrapping
    df["Сотрудники"] = df["Сотрудники"].astype(str).str.replace(", ", "\n")
    out_path = Path(tempfile.gettempdir()) / "Спектакли_и_кто_ведёт.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        sheet_name = "Спектакли"
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # Post-save auto-fit strictly by content (headers ignored)
    _post_save_autofit(out_path, sheet_name, min_width=10, max_width=80, wrap=True)

    return out_path


def spectacles_caption(total: int) -> str:
    return f"Спектакли и ответственные: {total} шт."