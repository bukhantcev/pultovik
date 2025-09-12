# services/ai_fill.py
from __future__ import annotations
from pathlib import Path
import base64
import csv
import io
import datetime as _dt
import calendar as _cal

from aiogram.types import FSInputFile
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from .prompt import SYSTEM_PROMPT
from config import OPENAI_API_KEY, GPT_MODEL

client = OpenAI(api_key=OPENAI_API_KEY)


async def describe_image(path: Path) -> str:
    """
    Отправляет изображение в OpenAI и возвращает его описание на русском языке.
    (Используется для отладочных сценариев, основная логика — build_excel_from_file / build_excel_from_site.)
    """
    ext = (path.suffix or '').lower()
    mime = 'image/jpeg'
    if ext == '.png':
        mime = 'image/png'
    elif ext == '.webp':
        mime = 'image/webp'

    img_bytes = Path(path).read_bytes()
    b64 = base64.b64encode(img_bytes).decode('ascii')
    data_url = f"data:{mime};base64,{b64}"

    resp = client.responses.create(
        model=GPT_MODEL,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": "Опиши подробно это изображение на русском языке."},
                {"type": "input_image", "image_url": data_url},
            ],
        }],
    )

    text = getattr(resp, 'output_text', None)
    return text if text else "Не удалось получить описание изображения."


# --- helpers and build_excel_from_file ---
CSV_HEADERS = ["Дата", "Тип", "Название", "Время", "Локация", "Город", "Сотрудник", "Дежурный сотрудник", "Инфо"]


def _to_data_url(path: Path) -> str:
    ext = (path.suffix or '').lower()
    mime = 'image/jpeg'
    if ext == '.png':
        mime = 'image/png'
    elif ext == '.webp':
        mime = 'image/webp'
    b64 = base64.b64encode(path.read_bytes()).decode('ascii')
    return f"data:{mime};base64,{b64}"


def _strip_code_fences(text: str) -> str:
    t = (text or '').strip()
    # remove common fences ```csv ... ``` or ``` ... ```
    if t.startswith('```'):
        parts = t.split('```')
        if len(parts) >= 3:
            body = parts[1]
            if body.strip().lower().startswith('csv'):
                body = body.split('\n', 1)[-1]
            return body.strip()
    return t


# --- ru date parsing & normalization ---
_RU_MONTHS_GEN = {
    1: "января", 2: "февраля", 3: "марта", 4: "апреля", 5: "мая", 6: "июня",
    7: "июля", 8: "августа", 9: "сентября", 10: "октября", 11: "ноября", 12: "декабря",
}
_RU_MONTH_PARSE = {
    "янв": 1, "января": 1, "январь": 1,
    "фев": 2, "февраля": 2, "февраль": 2,
    "мар": 3, "марта": 3, "март": 3,
    "апр": 4, "апреля": 4, "апрель": 4,
    "май": 5, "мая": 5,
    "июн": 6, "июня": 6, "июнь": 6,
    "июл": 7, "июля": 7, "июль": 7,
    "авг": 8, "августа": 8, "август": 8,
    "сен": 9, "сент": 9, "сентября": 9, "сентябрь": 9,
    "окт": 10, "октября": 10, "октябрь": 10,
    "ноя": 11, "ноября": 11, "ноябрь": 11,
    "дек": 12, "декабря": 12, "декабрь": 12,
}


def _parse_ru_date(s: str) -> _dt.date | None:
    s = (s or "").strip().lower()
    if not s:
        return None
    parts = [p for p in s.replace(',', ' ').split() if p]
    if len(parts) < 3:
        return None
    try:
        day = int(parts[0])
    except Exception:
        return None
    mon_str = parts[1]
    mon = _RU_MONTH_PARSE.get(mon_str[:3], _RU_MONTH_PARSE.get(mon_str))
    if not mon:
        return None
    year = None
    for p in parts[2:]:
        if p.isdigit() and len(p) >= 4:
            year = int(p)
            break
    if not year:
        return None
    try:
        return _dt.date(year, mon, day)
    except Exception:
        return None


def _human_ru_date(d: _dt.date) -> str:
    return f"{d.day} {_RU_MONTHS_GEN[d.month]} {d.year}"


def _normalize_csv(csv_text: str) -> list[list[str]]:
    f = io.StringIO(csv_text)
    reader = csv.reader(f)
    rows = [[c.strip() for c in r] for r in reader]
    if not rows:
        return [CSV_HEADERS]

    # Ensure headers, separate data rows
    if [h.strip() for h in rows[0]] == CSV_HEADERS:
        data_rows = rows[1:]
    else:
        data_rows = rows

    # Parse dates and collect
    parsed: list[tuple[_dt.date | None, list[str]]] = []
    dates_only: list[_dt.date] = []
    for r in data_rows:
        r = (r + [""] * len(CSV_HEADERS))[:len(CSV_HEADERS)]
        d = _parse_ru_date(r[0])
        parsed.append((d, r))
        if d:
            dates_only.append(d)

    if not dates_only:
        # вернуть как есть (с заголовком), если не распарсили ни одной даты
        return [CSV_HEADERS] + data_rows

    start = min(dates_only)
    month, year = start.month, start.year

    # Сводим по дате
    by_date: dict[_dt.date, list[list[str]]] = {}
    for d, r in parsed:
        if d:
            by_date.setdefault(d, []).append(r)
        else:
            # бездата — отправим в последний день месяца, чтобы не потерять
            last_day = _cal.monthrange(year, month)[1]
            dd = _dt.date(year, month, last_day)
            by_date.setdefault(dd, []).append(r)

    # Полный месяц с 1 по последний
    _, last_day = _cal.monthrange(year, month)
    ordered: list[list[str]] = []
    for day in range(1, last_day + 1):
        dd = _dt.date(year, month, day)
        if dd in by_date:
            # Записываем строки как есть; если дата пустая — оставляем дату как в входной строке
            for row in by_date[dd]:
                if not row[0]:
                    row[0] = _human_ru_date(dd)
                ordered.append(row)
        else:
            # Пустой день: создаём строку только с датой
            ordered.append([_human_ru_date(dd), "", "", "", "", "", "", "", ""])

    return [CSV_HEADERS] + ordered


def _rows_to_xlsx(rows: list[list[str]], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"

    for r in rows:
        # гарантируем длину
        r = (r + [""] * len(CSV_HEADERS))[:len(CSV_HEADERS)]
        ws.append(r)

    # стили заголовка
    if ws.max_row >= 1:
        bold = Font(bold=True)
        for cell in ws[1]:
            cell.font = bold
            cell.alignment = Alignment(vertical='center')

    # ширины столбцов
    widths = [14, 12, 34, 10, 18, 14, 18, 18, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(out_path)


async def build_excel_from_file(path: Path) -> FSInputFile:
    """
    Отправляет файл (фото/PDF/таблица) в GPT и собирает .xlsx по нашему шаблону.
    Гарантирует, что ВСЕ дни месяца присутствуют, а последняя строка не теряется.
    """
    instructions = (
        SYSTEM_PROMPT
        + "\n\nСтрого верни CSV с заголовками ровно в таком порядке: "
        + ", ".join(CSV_HEADERS)
        + ". Никаких комментариев, пояснений и блоков кода. Только CSV."
    )

    ext = (path.suffix or '').lower()
    if ext in {'.jpg', '.jpeg', '.png', '.webp'}:
        content = [
            {"type": "input_text", "text": instructions},
            {"type": "input_image", "image_url": _to_data_url(path)},
        ]
    else:
        in_file = client.files.create(file=open(path, 'rb'), purpose='assistants')
        content = [
            {"type": "input_text", "text": instructions},
            {"type": "input_file", "file_id": getattr(in_file, 'id', None)},
        ]

    resp = client.responses.create(
        model=GPT_MODEL,
        input=[{"role": "user", "content": content}],
    )

    text = getattr(resp, 'output_text', None) or str(resp)
    print(text)
    csv_text = _strip_code_fences(text)

    # Нормализуем: полное покрытие месяца + сохранение всех строк
    rows = _normalize_csv(csv_text)

    out_path = path.parent / f"ai_out_{path.stem}.xlsx"
    _rows_to_xlsx(rows, out_path)

    return FSInputFile(str(out_path))


async def build_excel_from_site(raw_text: str, month: int, year: int, tmp_name: str = "site") -> FSInputFile:
    """
    Принимает сырой текст, собранный с сайта (все элементы c-playbill--item),
    отправляет в GPT с инструкцией собрать CSV по шаблону, затем формирует .xlsx.

    raw_text — уже очищенный текст всего расписания за выбранный месяц.
    """
    # даём модели чёткие рамки: месяц/год уже известны
    prompt = (
        SYSTEM_PROMPT
        + "\n\n"
        + f"Работай строго для месяца: {month:02d}.{year} (все даты этого периода). "
          "Если в тексте отсутствуют некоторые дни, всё равно формируй строки для этих дней с пустыми полями кроме даты. "
          "Основная сцена - это Поварская, город Москва. "
        + "Строго верни CSV (UTF-8) с заголовками ровно в таком порядке: "
        + ", ".join(CSV_HEADERS)
        + ". Никаких комментариев и без блоков кода."
    )

    resp = client.responses.create(
        model=GPT_MODEL,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": raw_text},
            ],
        }],
    )

    text = getattr(resp, 'output_text', None) or str(resp)
    print(text)
    csv_text = _strip_code_fences(text)

    rows = _normalize_csv(csv_text)

    out_path = Path.cwd() / f"ai_out_{tmp_name}_{year}-{month:02d}.xlsx"
    _rows_to_xlsx(rows, out_path)
    return FSInputFile(str(out_path))